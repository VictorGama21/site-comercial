import streamlit as st
import psycopg2
import pandas as pd
from datetime import datetime, date, timedelta
import hashlib
from dateutil.relativedelta import relativedelta
import unidecode
import os
from dotenv import load_dotenv
import plotly.express as px
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://usuario:senha@host:porta/database")

# -----------------------------
# Segurança
# -----------------------------
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed


# -----------------------------
# Conexão / Init DB
# -----------------------------
def get_conn():
    return psycopg2.connect(DATABASE_URL)


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS stores (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            email TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('comercial','loja')),
            password_hash TEXT NOT NULL,
            store_id INTEGER REFERENCES stores(id)
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS visits (
            id SERIAL PRIMARY KEY,
            store_id INTEGER NOT NULL REFERENCES stores(id),
            visit_date DATE NOT NULL,
            weekday TEXT NOT NULL,
            buyer TEXT,
            supplier_id INTEGER REFERENCES suppliers(id),
            segment TEXT,
            warranty TEXT,
            info TEXT,
            status TEXT NOT NULL DEFAULT 'Pendente'
                CHECK(status IN ('Pendente','Concluída','Não Compareceu')),
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            completed_by INTEGER REFERENCES users(id),
            manager_comment TEXT
        );
    """)

    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS visits_unique_idx
        ON visits (store_id, visit_date, buyer, supplier_id, segment);
    """)

    conn.commit()
    conn.close()


# -----------------------------
# Utilitários
# -----------------------------
def get_stores():
    conn = get_conn()
    df = pd.read_sql_query("SELECT id, name FROM stores ORDER BY name;", conn)
    conn.close()
    return df


def get_suppliers():
    conn = get_conn()
    df = pd.read_sql_query("SELECT id, name FROM suppliers ORDER BY name;", conn)
    conn.close()
    return df


def ensure_supplier(name: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO suppliers(name)
        VALUES(%s)
        ON CONFLICT(name) DO UPDATE SET name=EXCLUDED.name
        RETURNING id;
    """, (name.strip(),))
    supplier_id = cur.fetchone()[0]
    conn.commit()
    conn.close()
    return supplier_id
# -----------------------------
# Configurações fixas
# -----------------------------
WEEKDAYS_PT = {
    0: "Segunda-feira", 1: "Terça-feira", 2: "Quarta-feira",
    3: "Quinta-feira", 4: "Sexta-feira", 5: "Sábado", 6: "Domingo",
}

SEGMENTOS_FIXOS = [
    "HORTIFRUTIGRANJEIRO", "EMBALAGEM", "CONGELADOS", "LATICINIOS", "SUPLEMENTOS",
    "PADARIA", "BEBIDAS", "MERCEARIA", "GRANJEIROS", "ACOUGUE", "OLEOS",
    "HIGIENE E BELEZA", "PET", "LIMPEZA DA CASA", "ECOMMERCE", "ROTISSERIA",
    "FRIOS E EMBUTIDOS", "QUEIJOS", "FLORICULTURA", "EMPORIO", "BAZAR"
]

ALLOWED_WARRANTY = {"", "Sim", "Não", "A confirmar"}


# -----------------------------
# Exportação Excel (colorido)
# -----------------------------
def export_visitas_excel(df):
    output = io.BytesIO()
    safe_df = df.copy()

    if "data_datetime" in safe_df.columns:
        safe_df = safe_df.drop(columns=["data_datetime"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False, sheet_name="Visitas")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    col_status = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "status":
            col_status = idx
            break

    if col_status:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=col_status, max_col=col_status):
            for cell in row:
                if cell.value and str(cell.value).lower() == "concluída":
                    cell.fill = PatternFill(start_color="C6EFCE",
                                            end_color="C6EFCE", fill_type="solid")
                elif cell.value and str(cell.value).lower() == "pendente":
                    cell.fill = PatternFill(start_color="FFC7CE",
                                            end_color="FFC7CE", fill_type="solid")
                elif cell.value and str(cell.value).lower() == "não compareceu":
                    cell.fill = PatternFill(start_color="FFD966",
                                            end_color="FFD966", fill_type="solid")

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()


# -----------------------------
# Página "Minhas Visitas" (Loja)
# -----------------------------
def page_minhas_visitas_loja():
    st.header("Minhas Visitas")
    user = st.session_state.user
    if not user or user["store_id"] is None:
        st.warning("Usuário não associado a nenhuma loja.")
        return

    hoje = date.today()
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    fim_semana = inicio_semana + timedelta(days=6)

    col4, col5 = st.columns(2)
    with col4:
        start = st.date_input("Início", value=inicio_semana, format="DD/MM/YYYY")
    with col5:
        end = st.date_input("Fim", value=fim_semana, format="DD/MM/YYYY")

    status = st.multiselect(
        "Status", ["Pendente", "Concluída", "Não Compareceu"],
        default=["Pendente", "Concluída", "Não Compareceu"]
    )

    dias_semana = ["Todos"] + list(WEEKDAYS_PT.values())
    dia_semana = st.selectbox("Filtrar por dia da semana", dias_semana)

    df = list_visits(store_id=user["store_id"], status=status, start=start, end=end)
    if dia_semana != "Todos":
        df = df[df["dia_semana"] == dia_semana]

    if df.empty:
        st.info("Nenhuma visita encontrada para os filtros selecionados.")
        return

    df["data_datetime"] = pd.to_datetime(df["data"], format="%d/%m/%Y")
    hoje_ts = pd.Timestamp(date.today())
    pendentes_vencidas = df[(df["status"] == "Pendente") & (df["data_datetime"] < hoje_ts)]

    if not pendentes_vencidas.empty:
        st.warning(f"⚠️ Existem {len(pendentes_vencidas)} visita(s) pendente(s) com data anterior a hoje!")

    excel_bytes = export_visitas_excel(df)
    st.download_button(
        "📥 Baixar visitas (Excel)", data=excel_bytes,
        file_name="minhas_visitas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.metric("Total de visitas", len(df))
    st.metric("Concluídas", (df["status"] == "Concluída").sum())
    st.metric("Não Compareceu", (df["status"] == "Não Compareceu").sum())

    st.subheader("📋 Lista de Visitas")
    for _, row in df.iterrows():
        with st.container():
            col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 1])
            col1.write(f"📅 **Data:** {row['data']}")
            col2.write(f"📆 **Dia da semana:** {row['dia_semana']}")
            col3.write(f"👤 **Comprador:** {row['comprador']}")
            col4.write(f"🏢 **Fornecedor:** {row['fornecedor']}")
            col5.write(f"📦 **Segmento:** {row['segmento']}")

            col1, col2, col3 = st.columns([2, 2, 4])
            col1.write(f"🛡 **Garantia:** {row['garantia']}")
            col2.write(f"📌 **Status:** {row['status']}")
            col3.write(f"📝 **Info:** {row['info'] if row['info'] else '-'}")

            if row.get("manager_comment"):
                st.info(f"💬 **Comentário do Gerente:** {row['manager_comment']}")

            if row["status"] == "Pendente":
                comentario = st.text_area("💬 Observação (opcional)", key=f"comentario_{row['id']}")
                colA, colB = st.columns(2)
                with colA:
                    if st.button("✅ Concluir", key=f"concluir_{row['id']}"):
                        concluir_visit(row["id"], user["id"], comentario if comentario.strip() else None)
                        st.success(f"Visita {row['id']} concluída com sucesso!")
                        st.rerun()
                with colB:
                    if st.button("❌ Fornecedor não foi", key=f"nao_compareceu_{row['id']}"):
                        nao_compareceu_visit(row["id"], user["id"], comentario if comentario.strip() else None)
                        st.warning(f"Visita {row['id']} marcada como 'Não Compareceu'.")
                        st.rerun()
            elif row["status"] in ["Concluída", "Não Compareceu"]:
                if row["status"] == "Concluída":
                    st.write("✔️ **Visita concluída**")
                elif row["status"] == "Não Compareceu":
                    st.write("⚠️ **Promotor não compareceu**")

                if st.button("🔄 Reabrir visita", key=f"reabrir_{row['id']}"):
                    reabrir_visit(row["id"], user["id"])
                    st.info(f"Visita {row['id']} reaberta e agora está Pendente.")
                    st.rerun()

            st.markdown("---")

    with st.expander("❓ Precisa de ajuda?"):
        st.markdown("""
        Caso esteja com dúvidas ou problemas com a agenda de visitas, entre em contato com o setor de compras:
        📧 **Email:** [compras1@quitandaria.com.br](mailto:compras1@quitandaria.com.br)
        """)
# -----------------------------
# Listagem e atualização de visitas
# -----------------------------
def list_visits(store_id=None, status=None, start=None, end=None):
    conn = get_conn()
    cur = conn.cursor()

    q = [
        "SELECT v.id, s.name AS loja, to_char(v.visit_date, 'DD/MM/YYYY') AS data,",
        "v.weekday AS dia_semana, v.buyer AS comprador, sp.name AS fornecedor,",
        "v.segment AS segmento, v.warranty AS garantia, v.info AS info,",
        "v.status, v.manager_comment",
        "FROM visits v",
        "JOIN stores s ON s.id = v.store_id",
        "JOIN suppliers sp ON sp.id = v.supplier_id",
        "WHERE 1=1"
    ]
    params = []

    if store_id:
        q.append("AND v.store_id = %s")
        params.append(store_id)

    if status:
        if len(status) == 1:
            q.append("AND v.status = %s")
            params.append(status[0])
        else:
            q.append("AND v.status = ANY(%s)")
            params.append(status)

    if start:
        q.append("AND v.visit_date >= %s")
        params.append(start)
    if end:
        q.append("AND v.visit_date <= %s")
        params.append(end)

    q.append("ORDER BY v.visit_date DESC")
    cur.execute(" ".join(q), tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    cols = [
        "id", "loja", "data", "dia_semana", "comprador", "fornecedor",
        "segmento", "garantia", "info", "status", "manager_comment"
    ]
    return pd.DataFrame(rows, columns=cols)


def update_visit(visit_id: int, buyer: str, supplier: str,
                 segment: str, warranty: str, info: str):
    supplier_id = ensure_supplier(supplier)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE visits
        SET buyer=%s, supplier_id=%s, segment=%s,
            warranty=%s, info=%s
        WHERE id=%s;
    """, (buyer, supplier_id, segment, warranty, info, visit_id))
    conn.commit()
    conn.close()


def delete_visit(visit_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM visits WHERE id=%s;", (visit_id,))
    conn.commit()
    conn.close()


# -----------------------------
# Estilo de tabela
# -----------------------------
def highlight_status(val):
    if val == "Concluída":
        return "background-color: #90EE90; color: black;"
    elif val == "Pendente":
        return "background-color: #FF7F7F; color: black;"
    elif val == "Não Compareceu":
        return "background-color: #FFD966; color: black;"
    return ""


def style_table(df: pd.DataFrame):
    return df.style.applymap(highlight_status, subset=["status"])


# -----------------------------
# Dashboard Comercial
# -----------------------------
def page_dashboard_comercial():
    st.header("Agenda Geral")

    stores = get_stores()
    stores_filter = ["Todas"] + stores["name"].tolist()
    dias_semana = ["Todos"] + list(WEEKDAYS_PT.values())

    col1, col2, col3 = st.columns(3)
    with col1:
        loja_nome = st.selectbox("Loja", stores_filter)
        loja_id = None if loja_nome == "Todas" else int(stores.loc[stores["name"] == loja_nome, "id"].iloc[0])
    with col2:
        status = st.multiselect("Status",
                                ["Pendente", "Concluída", "Não Compareceu"],
                                default=["Pendente", "Concluída", "Não Compareceu"])
    with col3:
        dia_semana = st.selectbox("Dia da semana", dias_semana)

    hoje = date.today()
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    fim_semana = inicio_semana + timedelta(days=6)

    col4, col5 = st.columns(2)
    with col4:
        start = st.date_input("Início", value=inicio_semana, format="DD/MM/YYYY")
    with col5:
        end = st.date_input("Fim", value=fim_semana, format="DD/MM/YYYY")

    df = list_visits(store_id=loja_id, status=status, start=start, end=end)
    if dia_semana != "Todos":
        df = df[df["dia_semana"] == dia_semana]

    if df.empty:
        st.info("Sem visitas no período ou nos filtros selecionados.")
        return

    st.dataframe(style_table(df), use_container_width=True, hide_index=True)

    st.metric("Total de visitas", len(df))
    st.metric("Concluídas", (df["status"] == "Concluída").sum())

    st.subheader("📊 Dashboard Analítico")
    col1, col2 = st.columns(2)
    with col1:
        fig1 = px.histogram(df, x="segmento", color="status", title="Visitas por Segmento")
        st.plotly_chart(fig1, use_container_width=True)
    with col2:
        fig2 = px.histogram(df, x="loja", color="status", title="Visitas por Loja")
        st.plotly_chart(fig2, use_container_width=True)

    fig3 = px.line(df, x="data", color="status", title="Evolução das Visitas")
    st.plotly_chart(fig3, use_container_width=True)

    st.subheader("✏️ Gerenciar Visitas")
    visit_id = st.selectbox("Selecione uma visita", df["id"].tolist())
    if visit_id:
        vrow = df[df["id"] == visit_id].iloc[0]

        comprador = st.text_input("Comprador", vrow["comprador"])
        fornecedor = st.text_input("Fornecedor", vrow["fornecedor"])
        segmento = st.selectbox("Segmento", SEGMENTOS_FIXOS,
                                index=SEGMENTOS_FIXOS.index(vrow["segmento"])
                                if vrow["segmento"] in SEGMENTOS_FIXOS else 0)
        garantia = st.selectbox("Garantia", ["", "Sim", "Não", "A confirmar"],
                                index=["", "Sim", "Não", "A confirmar"].index(vrow["garantia"])
                                if vrow["garantia"] in ["", "Sim", "Não", "A confirmar"] else 0)
        info = st.text_area("Informações", vrow["info"])

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            if st.button("Salvar alterações"):
                update_visit(visit_id, comprador, fornecedor, segmento, garantia, info)
                st.success("Visita atualizada!")
                st.rerun()
        with col2:
            if st.button("Excluir visita"):
                delete_visit(visit_id)
                st.warning("Visita excluída!")
                st.rerun()
        with col3:
            if vrow["status"] == "Concluída":
                if st.button("🔄 Reabrir visita"):
                    reabrir_visit(visit_id, st.session_state.user["id"])
                    st.info("Visita reaberta e agora está Pendente.")
                    st.rerun()
        with col4:
            if vrow["status"] == "Pendente":
                comentario = st.text_area("💬 Comentário do Gerente (opcional)", key=f"comentario_{visit_id}")
                if st.button("✅ Concluir visita", key=f"concluir_{visit_id}"):
                    concluir_visit(visit_id, st.session_state.user["id"], comentario if comentario.strip() else None)
                    st.success("Visita concluída!")
                    st.rerun()


# -----------------------------
# Login / Logout
# -----------------------------
def login_form():
    st.title("Login - Sistema de Visitas")
    email = st.text_input("Email")
    password = st.text_input("Senha", type="password")

    if st.button("Entrar"):
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, email, name, role, password_hash, store_id FROM users WHERE email = %s;", (email,))
        user = cur.fetchone()
        conn.close()

        if user and verify_password(password, user[4]):
            st.session_state.user = {
                "id": user[0], "email": user[1], "name": user[2],
                "role": user[3], "store_id": user[5]
            }
            st.success(f"Bem-vindo(a), {user[2]}!")
            st.rerun()
        else:
            st.error("Email ou senha incorretos.")


def logout_button():
    if st.sidebar.button("Sair"):
        st.session_state.user = None
        st.rerun()


# -----------------------------
# Rodapé
# -----------------------------
def footer():
    st.markdown(
        """
        ---
        <div style='text-align: center; font-size: 12px; color: gray; line-height: 1.6;'>
        📱 <b>Sistema de Visitas - Quitandaria</b><br>
        © 2025 Victor Manuel Gama dos Anjos – Todos os direitos reservados<br><br>
        🚀 Desenvolvido para facilitar a gestão de visitas e fornecedores.<br><br>
        🔒 Informações protegidas · 📦 Sujeito à disponibilidade<br><br>
        Bairro Novo – Av. Presidente Getúlio Vargas, 761, Olinda - PE
        </div>
        """, unsafe_allow_html=True
    )
    
def seed_data():
    conn = get_conn()
    cur = conn.cursor()

    lojas = [
        "HIPODROMO", "RIO DOCE", "CARUARU", "HIPODROMO CAFETERIA", "JANGA CAFETERIA",
        "ESPINHEIRO", "AFLITOS", "PONTA VERDE", "JATIUCA", "FAROL", "BEIRA MAR",
        "JARDIM ATLÂNTICO", "CASA CAIADA VERDAO", "JANGA VERDAO", "BAIRRO NOVO VERDAO"
    ]

    # cria lojas se não existirem
    cur.execute("SELECT COUNT(*) FROM stores;")
    if cur.fetchone()[0] == 0:
        for loja in lojas:
            cur.execute("INSERT INTO stores(name) VALUES(%s) ON CONFLICT DO NOTHING;", (loja,))

    cur.execute("SELECT id, name FROM stores;")
    stores_map = {name: _id for _id, name in cur.fetchall()}

    # cria usuários se não existirem
    cur.execute("SELECT COUNT(*) FROM users;")
    if cur.fetchone()[0] == 0:
        users = [
            ("comercial@quitandaria.com", "Comercial Master", "comercial",
             hash_password("123456"), None)
        ]
        for loja in lojas:
            email_loja = "loja." + unidecode.unidecode(loja.lower().replace(" ", ".")) + "@quitandaria.com"
            users.append((email_loja, loja, "loja", hash_password("123456"), stores_map.get(loja)))

        cur.executemany(
            "INSERT INTO users(email, name, role, password_hash, store_id) VALUES(%s,%s,%s,%s,%s);",
            users
        )

    conn.commit()
    conn.close()


# -----------------------------
# Main App
# -----------------------------
def main():
    st.set_page_config(page_title="Sistema de Visitas", layout="wide")
    st.markdown(
        """
        <style>
        body {background: linear-gradient(135deg, #FF8C42, #FFB347);}
        .stButton>button {background-color: #FF8C42; color: white; font-weight: bold; border-radius: 10px;}
        </style>
        """, unsafe_allow_html=True
    )

    init_db()
    seed_data()

    st.sidebar.title("📅 Sistema de Visitas - Quitandaria")

    if "user" not in st.session_state:
        st.session_state.user = None

    if st.session_state.user is None:
        login_form()
        footer()
        return

    user = st.session_state.user
    st.sidebar.write(f"**Usuário:** {user['name']}")
    st.sidebar.write(f"**Perfil:** {user['role'].capitalize()}")

    if user["role"] == "comercial":
        page = st.sidebar.radio("Navegação", ["Agenda Geral", "Agendar Visita"])
        logout_button()
        if page == "Agenda Geral":
            page_dashboard_comercial()
        else:
            page_agendar_visita()
        footer()
    elif user["role"] == "loja":
        st.sidebar.radio("Navegação", ["Minhas Visitas"])
        logout_button()
        page_minhas_visitas_loja()
        footer()


if __name__ == "__main__":
    main()

